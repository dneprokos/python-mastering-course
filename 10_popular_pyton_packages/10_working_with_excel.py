# python .\10_popular_pyton_packages\10_working_with_excel.py

import openpyxl
from pathlib import Path

path = Path("10_popular_pyton_packages") / "transactions.xlsx"

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(path)
print(wb.sheetnames)

sheet = wb["Sheet1"]  # ['Sheet1']
wb.create_sheet("Sheet2", 0)  # will create a sheet before
print(wb.sheetnames)  # ['Sheet2', 'Sheet1']
del wb["Sheet2"]  # Will delete a sheet
print(wb.sheetnames)  # ['Sheet1']

cell = sheet["a1"]
print(cell.value)  # transaction_id
print(cell.row)  # 1
print(cell.column)  # 1
print(cell.coordinate)  # A1

# alternative way to access a cell
cell = sheet.cell(row=1, column=1)
print(cell.coordinate)  # A1

print(sheet.max_row)  # 4
print(sheet.max_column)  # 3

print("\nPrint the value of each cell")
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

# transaction_id
# product_id
# price
# 1001
# 1
# 5.95
# 1002
# 2
# 6.95
# 1003
# 3
# 7.95

column = sheet["a"]
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>)
print(column)

cells = sheet["a:c"]
print(cells)  # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>), (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>), (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.C4>))

cells = sheet["a1:c3"]
print(cells)  # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))

# sheet.append([1, 2, 3])
